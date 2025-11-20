"""Benchmark script for comparing Torch and OpenVINO inference performance.

This script trains a model, exports it to Torch and multiple OpenVINO formats,
and benchmarks inference performance on all formats (PyTorch, OpenVINO FP32, FP16, INT8).

Usage:
    python torch_ov_difference.py --device cuda --num-inferences 100
    python torch_ov_difference.py --device cpu --num-inferences 50
"""

import argparse
import os
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from anomalib.data import MVTecAD
from anomalib.deploy import CompressionType, OpenVINOInferencer, TorchInferencer
from anomalib.engine import Engine
from anomalib.models import get_model

os.environ["TRUST_REMOTE_CODE"] = "1"
os.environ["CUDA_VISIBLE_DEVICES"] = "0"


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Benchmark Torch vs OpenVINO inference performance (FP32, FP16, INT8)",
    )
    parser.add_argument(
        "--device",
        type=str,
        default="cuda",
        choices=["cuda", "cpu", "CPU", "GPU"],
        help="Device to run inference on (cuda/cpu for Torch, CPU/GPU for OpenVINO)",
    )
    parser.add_argument(
        "--model",
        type=str,
        nargs="+",
        default=["Padim"],
        help="Model(s) to use for benchmarking. Can specify multiple models separated by spaces (default: Padim)",
    )
    parser.add_argument(
        "--num-inferences",
        type=int,
        default=100,
        help="Number of inference runs for benchmarking (default: 100)",
    )
    parser.add_argument(
        "--category",
        type=str,
        default="transistor",
        help="MVTec AD category to use (default: transistor)",
    )
    parser.add_argument(
        "--skip-training",
        action="store_true",
        help="Skip training and use existing exported models",
    )
    parser.add_argument(
        "--model-path",
        type=str,
        default=None,
        help="Path to existing model directory (used with --skip-training)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output Excel file path (default: benchmark_results_YYYYMMDD_HHMMSS.xlsx)",
    )
    return parser.parse_args()


def save_results_to_excel(all_models_results, config, output_path):
    """Save benchmark results to an Excel file.

    Args:
        all_models_results: Dictionary mapping model names to their results dict
        config: Dictionary with configuration parameters
        output_path: Path to save the Excel file
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create a summary sheet first
    summary_sheet = wb.create_sheet("Summary", 0)

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)

    # ===== SUMMARY SHEET =====
    row = 1
    summary_sheet.merge_cells(f"A{row}:G{row}")
    cell = summary_sheet[f"A{row}"]
    cell.value = "Multi-Model Inference Benchmark Results - Summary"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # Configuration section
    summary_sheet[f"A{row}"] = "Configuration"
    summary_sheet[f"A{row}"].font = subheader_font
    summary_sheet[f"A{row}"].fill = subheader_fill
    row += 1

    config_items = [
        ("Models", ", ".join(config.get("models", ["N/A"]))),
        ("Category", config.get("category", "N/A")),
        ("Device", config.get("device", "N/A")),
        ("Number of Inferences", config.get("num_inferences", "N/A")),
        ("Test Images", config.get("test_images", "N/A")),
        ("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for key, value in config_items:
        summary_sheet[f"A{row}"] = key
        summary_sheet[f"B{row}"] = value
        row += 1

    row += 2

    # Summary comparison table across all models
    summary_sheet[f"A{row}"] = "Cross-Model Performance Summary"
    summary_sheet[f"A{row}"].font = subheader_font
    summary_sheet[f"A{row}"].fill = subheader_fill
    row += 1

    # Headers for summary table
    headers = ["Model", "Format", "Avg Time (s)", "FPS", "Min Time (s)", "Max Time (s)", "Export Time (s)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Fill summary data
    for model_name, model_results in all_models_results.items():
        for format_name in ["PyTorch", "OpenVINO (FP32)", "OpenVINO (FP16)", "OpenVINO (INT8)"]:
            if format_name in model_results:
                result = model_results[format_name]
                summary_sheet[f"A{row}"] = model_name
                summary_sheet[f"B{row}"] = format_name
                summary_sheet[f"C{row}"] = round(result["avg_time"], 4)
                summary_sheet[f"D{row}"] = round(result["fps"], 2)
                summary_sheet[f"E{row}"] = round(result["min_time"], 4)
                summary_sheet[f"F{row}"] = round(result["max_time"], 4)
                summary_sheet[f"G{row}"] = round(result["export_time"], 2) if "export_time" in result else "N/A"
                row += 1

    # Adjust column widths for summary
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        summary_sheet.column_dimensions[col].width = 18

    # ===== CREATE INDIVIDUAL SHEETS FOR EACH MODEL =====
    for model_name, model_results in all_models_results.items():
        ws = wb.create_sheet(model_name)

        # Title
        row = 1
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        # Title
        row = 1
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = f"Model Inference Benchmark Results - {model_name}"
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center")
        row += 2

        # Configuration section
        ws[f"A{row}"] = "Configuration"
        ws[f"A{row}"].font = subheader_font
        ws[f"A{row}"].fill = subheader_fill
        row += 1

        model_config_items = [
            ("Model", model_name),
            ("Category", config.get("category", "N/A")),
            ("Device", config.get("device", "N/A")),
            ("Number of Inferences", config.get("num_inferences", "N/A")),
            ("Test Images", config.get("test_images", "N/A")),
            ("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for key, value in model_config_items:
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value
            row += 1

        row += 1

        # Results section
        ws[f"A{row}"] = "Benchmark Results"
        ws[f"A{row}"].font = subheader_font
        ws[f"A{row}"].fill = subheader_fill
        row += 1

        # Column headers
        headers = [
            "Model Type",
            "Total Time (s)",
            "Avg Time (s)",
            "Min Time (s)",
            "Max Time (s)",
            "FPS",
            "Export Time (s)",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Data rows
        model_order = ["PyTorch", "OpenVINO (FP32)", "OpenVINO (FP16)", "OpenVINO (INT8)"]
        for format_name in model_order:
            if format_name in model_results:
                result = model_results[format_name]
                ws[f"A{row}"] = format_name
                ws[f"B{row}"] = round(result["total_time"], 4)
                ws[f"C{row}"] = round(result["avg_time"], 4)
                ws[f"D{row}"] = round(result["min_time"], 4)
                ws[f"E{row}"] = round(result["max_time"], 4)
                ws[f"F{row}"] = round(result["fps"], 2)
                ws[f"G{row}"] = round(result["export_time"], 2) if "export_time" in result else "N/A"
                row += 1

        row += 1

        # Speedup comparison section
        ws[f"A{row}"] = "Speedup vs PyTorch"
        ws[f"A{row}"].font = subheader_font
        ws[f"A{row}"].fill = subheader_fill
        row += 1

        # Speedup headers
        speedup_headers = ["Model Type", "Avg Time Speedup", "FPS Speedup"]
        for col_idx, header in enumerate(speedup_headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Calculate speedups
        if "PyTorch" in model_results:
            torch_results = model_results["PyTorch"]
            for format_name in ["OpenVINO (FP32)", "OpenVINO (FP16)", "OpenVINO (INT8)"]:
                if format_name in model_results:
                    result = model_results[format_name]
                    avg_speedup = torch_results["avg_time"] / result["avg_time"] if result["avg_time"] > 0 else 0
                    fps_speedup = result["fps"] / torch_results["fps"] if torch_results["fps"] > 0 else 0

                    ws[f"A{row}"] = format_name
                    ws[f"B{row}"] = f"{avg_speedup:.2f}x"
                    ws[f"C{row}"] = f"{fps_speedup:.2f}x"
                    row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 18

    # Save workbook
    wb.save(output_path)
    print(f"\nResults saved to: {output_path}")
    print(f"Excel file contains {len(all_models_results)} model sheet(s) plus a summary sheet")


def benchmark_inferencer(inferencer, data, num_inferences, inferencer_name, device):
    """Run benchmark for a given inferencer.

    Args:
        inferencer: The inferencer instance (TorchInferencer or OpenVINOInferencer)
        data: Data module containing test data
        num_inferences: Number of inferences to run
        inferencer_name: Name for logging (e.g., "TorchInferencer")
        device: Device being used

    Returns:
        dict: Benchmark results including total time, avg time, and FPS
    """
    """Run benchmark for a given inferencer.

    Args:
        inferencer: The inferencer instance (TorchInferencer or OpenVINOInferencer)
        data: Data module containing test data
        num_inferences: Number of inferences to run
        inferencer_name: Name for logging (e.g., "TorchInferencer")
        device: Device being used

    Returns:
        dict: Benchmark results including total time, avg time, and FPS
    """
    print(f"\n{'=' * 60}")
    print(f"Benchmarking {inferencer_name} on {device}")
    print(f"{'=' * 60}")

    # Validate test data exists
    if not hasattr(data, "test_data") or len(data.test_data) == 0:
        raise ValueError("No test data available. Please check the dataset.")

    timings = []

    # Warmup run (exclude from timing)
    sample = data.test_data[0]
    _ = inferencer.predict(sample.image)
    print("Warmup complete. Starting benchmark...")

    # Benchmark runs
    for i in range(num_inferences):
        sample = data.test_data[i % len(data.test_data)]

        tic = time.time()
        result = inferencer.predict(sample.image)
        toc = time.time()

        inference_time = toc - tic
        timings.append(inference_time)

        if (i + 1) % 10 == 0 or i == 0:
            print(f"  [{i + 1}/{num_inferences}] Inference time: {inference_time:.4f}s")

    # Calculate statistics
    total_time = sum(timings)
    avg_time = total_time / num_inferences
    fps = num_inferences / total_time
    min_time = min(timings)
    max_time = max(timings)

    results = {
        "total_time": total_time,
        "avg_time": avg_time,
        "fps": fps,
        "min_time": min_time,
        "max_time": max_time,
        "num_inferences": num_inferences,
    }

    print(f"\n{inferencer_name} Results:")
    print(f"  Total time:        {total_time:.4f}s")
    print(f"  Average time:      {avg_time:.4f}s")
    print(f"  Min time:          {min_time:.4f}s")
    print(f"  Max time:          {max_time:.4f}s")
    print(f"  FPS:               {fps:.2f}")

    return results


def main():
    """Main function to run the benchmark."""
    args = parse_args()

    # Convert single model to list if necessary
    models = args.model if isinstance(args.model, list) else [args.model]

    print(f"\n{'=' * 60}")
    print("Multi-Model Inference Benchmark Configuration")
    print(f"{'=' * 60}")
    print(f"Models:            {', '.join(models)}")
    print(f"Device:            {args.device}")
    print(f"Number of runs:    {args.num_inferences}")
    print(f"Category:          {args.category}")
    print(f"Skip training:     {args.skip_training}")
    print("Comparing:         PyTorch, OpenVINO (FP32), OpenVINO (FP16), OpenVINO (INT8)")
    print(f"{'=' * 60}\n")

    # Load data (shared across all models)
    print("Loading dataset...")
    data = MVTecAD(category=args.category, num_workers=0)
    data.setup()  # Initialize data splits

    # Dictionary to store all models' results
    all_models_results = {}

    # Map device names
    # Torch uses "cuda"/"cpu", OpenVINO uses "GPU"/"CPU"
    device_lower = args.device.lower()
    if device_lower == "cuda":
        torch_device = "cuda"
        ov_device = "GPU"
    elif device_lower == "cpu":
        torch_device = "cpu"
        ov_device = "CPU"
    elif device_lower == "gpu":
        torch_device = "cuda"  # Map GPU to cuda for Torch
        ov_device = "GPU"
    else:
        # Default to CPU for both
        torch_device = "cpu"
        ov_device = "CPU"

    # Process each model
    for model_idx, model_name in enumerate(models):
        print(f"\n{'#' * 60}")
        print(f"# Processing Model {model_idx + 1}/{len(models)}: {model_name}")
        print(f"{'#' * 60}\n")

        # Dictionary to store results for this model
        model_results = {}

        # Get model paths
        if args.skip_training and args.model_path:
            torch_path = Path(args.model_path) / f"{model_name}_model.pt"
            ov_fp32_path = Path(args.model_path) / f"{model_name}_model.xml"
            ov_fp16_path = Path(args.model_path) / f"{model_name}_model_fp16.xml"
            ov_int8_path = Path(args.model_path) / f"{model_name}_model_int8_ptq.xml"

            # Validate files exist
            if not torch_path.exists():
                print(f"Warning: Torch model not found: {torch_path}. Skipping {model_name}.")
                continue
            if not ov_fp32_path.exists():
                print(f"Warning: OpenVINO FP32 model not found: {ov_fp32_path}. Skipping {model_name}.")
                continue

            print(f"\nUsing existing models from: {args.model_path}")
        else:
            # Train and export model
            print(f"\nInitializing model {model_name} and training...")
            try:
                model = get_model(model_name)
            except Exception as e:
                print(f"Error initializing model {model_name}: {e}")
                print(f"Skipping {model_name}.")
                continue

            engine = Engine(max_epochs=5)

            print(f"Training {model_name}...")
            try:
                engine.fit(datamodule=data, model=model)
            except Exception as e:
                print(f"Error training model {model_name}: {e}")
                print(f"Skipping {model_name}.")
                continue

            print(f"Testing {model_name}...")
            try:
                engine.test(datamodule=data, model=model)
            except Exception as e:
                print(f"Warning: Error testing model {model_name}: {e}")
                print("Continuing with export...")

            # Export models
            print(f"\nExporting {model_name} models...")
            filename = f"{model_name}_{args.category}_benchmark"

            # Export PyTorch model
            print(f"Exporting {model_name} PyTorch model...")
            try:
                torch_path = engine.export(model=model, export_type="torch", model_file_name=filename)
                print(f"PyTorch model exported to: {torch_path}")
            except Exception as e:
                print(f"Error exporting PyTorch model for {model_name}: {e}")
                print(f"Skipping {model_name}.")
                continue

            # Export OpenVINO FP32 model
            print(f"Exporting {model_name} OpenVINO FP32 model...")
            try:
                fp32_export_start = time.time()
                ov_fp32_path = engine.export(model=model, export_type="openvino", model_file_name=filename)
                fp32_export_time = time.time() - fp32_export_start
                print(f"OpenVINO FP32 model exported to: {ov_fp32_path}")
                print(f"FP32 export time: {fp32_export_time:.2f}s")
            except Exception as e:
                print(f"Error exporting OpenVINO FP32 model for {model_name}: {e}")
                print(f"Skipping {model_name}.")
                continue

            # Export OpenVINO FP16 model
            print(f"Exporting {model_name} OpenVINO FP16 model...")
            try:
                fp16_export_start = time.time()
                ov_fp16_path = engine.export(
                    model=model,
                    export_type="openvino",
                    model_file_name=f"{filename}_fp16",
                    compression_type=CompressionType.FP16,
                )
                fp16_export_time = time.time() - fp16_export_start
                print(f"OpenVINO FP16 model exported to: {ov_fp16_path}")
                print(f"FP16 export time: {fp16_export_time:.2f}s")
            except Exception as e:
                print(f"Warning: Failed to export FP16 model for {model_name}: {e}")
                ov_fp16_path = None
                fp16_export_time = None

            # Export OpenVINO INT8 model
            print(f"Exporting {model_name} OpenVINO INT8 model...")
            try:
                int8_export_start = time.time()
                ov_int8_path = engine.export(
                    model=model,
                    export_type="openvino",
                    model_file_name=f"{filename}_int8_ptq",
                    compression_type=CompressionType.INT8_PTQ,
                    datamodule=data,
                )
                int8_export_time = time.time() - int8_export_start
                print(f"OpenVINO INT8 model exported to: {ov_int8_path}")
                print(f"INT8 export time: {int8_export_time:.2f}s")
            except Exception as e:
                print(f"\nWarning: Failed to export INT8 model for {model_name}: {e}")
                if "'ImageBatch' object is not subscriptable" in str(e):
                    print("Note: This is a known issue with NNCF and ImageBatch format.")
                    print("Workaround: The anomalib export_mixin.py needs to handle ImageBatch objects.")
                print(f"Continuing with PyTorch, FP32 and FP16 models only for {model_name}...\n")
                ov_int8_path = None
                int8_export_time = None

        # Benchmark PyTorch Inferencer
        print("\n" + "=" * 60)
        print(f"Starting {model_name} PyTorch Inference Benchmark")
        print("=" * 60)
        try:
            torch_inferencer = TorchInferencer(path=torch_path, device=torch_device)
            torch_results = benchmark_inferencer(
                torch_inferencer,
                data,
                args.num_inferences,
                f"{model_name} PyTorch",
                torch_device,
            )
            model_results["PyTorch"] = torch_results
        except Exception as e:
            print(f"Error benchmarking PyTorch inferencer for {model_name}: {e}")

        # Benchmark OpenVINO FP32 Inferencer
        print("\n" + "=" * 60)
        print(f"Starting {model_name} OpenVINO (FP32) Inference Benchmark")
        print("=" * 60)
        try:
            ov_fp32_inferencer = OpenVINOInferencer(path=ov_fp32_path, device=ov_device)
            ov_fp32_results = benchmark_inferencer(
                ov_fp32_inferencer,
                data,
                args.num_inferences,
                f"{model_name} OpenVINO (FP32)",
                ov_device,
            )
            if not args.skip_training:
                ov_fp32_results["export_time"] = fp32_export_time
            model_results["OpenVINO (FP32)"] = ov_fp32_results
        except Exception as e:
            print(f"Error benchmarking OpenVINO FP32 inferencer for {model_name}: {e}")

        # Benchmark OpenVINO FP16 Inferencer
        if ov_fp16_path and Path(ov_fp16_path).exists():
            print("\n" + "=" * 60)
            print(f"Starting {model_name} OpenVINO (FP16) Inference Benchmark")
            print("=" * 60)
            try:
                ov_fp16_inferencer = OpenVINOInferencer(path=ov_fp16_path, device=ov_device)
                ov_fp16_results = benchmark_inferencer(
                    ov_fp16_inferencer,
                    data,
                    args.num_inferences,
                    f"{model_name} OpenVINO (FP16)",
                    ov_device,
                )
                if not args.skip_training and fp16_export_time is not None:
                    ov_fp16_results["export_time"] = fp16_export_time
                model_results["OpenVINO (FP16)"] = ov_fp16_results
            except Exception as e:
                print(f"Error benchmarking OpenVINO FP16 inferencer for {model_name}: {e}")
        else:
            print(f"\nSkipping OpenVINO (FP16) benchmark for {model_name} (model not available)")

        # Benchmark OpenVINO INT8 Inferencer (if export was successful)
        if ov_int8_path and Path(ov_int8_path).exists():
            print("\n" + "=" * 60)
            print(f"Starting {model_name} OpenVINO (INT8) Inference Benchmark")
            print("=" * 60)
            try:
                ov_int8_inferencer = OpenVINOInferencer(path=ov_int8_path, device=ov_device)
                ov_int8_results = benchmark_inferencer(
                    ov_int8_inferencer,
                    data,
                    args.num_inferences,
                    f"{model_name} OpenVINO (INT8)",
                    ov_device,
                )
                if not args.skip_training and int8_export_time is not None:
                    ov_int8_results["export_time"] = int8_export_time
                model_results["OpenVINO (INT8)"] = ov_int8_results
            except Exception as e:
                print(f"Error benchmarking OpenVINO INT8 inferencer for {model_name}: {e}")
        else:
            print("\n" + "=" * 60)
            print(f"Skipping {model_name} OpenVINO (INT8) benchmark (export failed or not available)")
            print("=" * 60)

        # Store results for this model
        if model_results:
            all_models_results[model_name] = model_results

            # Print summary for this model
            print("\n" + "=" * 60)
            print(f"BENCHMARK SUMMARY FOR {model_name}")
            print("=" * 60)
            print(f"Category:          {args.category}")
            print(f"Device:            Torch={torch_device.upper()}, OpenVINO={ov_device}")
            print(f"Test Images:       {len(data.test_data)}")
            print(f"Num Inferences:    {args.num_inferences}")
            print("=" * 60)
            print(f"{'Format':<20} {'Avg Time (s)':<15} {'FPS':<15}")
            print("-" * 60)

            for format_name in ["PyTorch", "OpenVINO (FP32)", "OpenVINO (FP16)", "OpenVINO (INT8)"]:
                if format_name in model_results:
                    result = model_results[format_name]
                    print(f"{format_name:<20} {result['avg_time']:<15.4f} {result['fps']:<15.2f}")

            print("=" * 60)
            print(f"{'Format':<20} {'Avg Speedup':<15} {'FPS Speedup':<15}")
            print("-" * 60)

            # Calculate speedups vs PyTorch for this model
            if "PyTorch" in model_results:
                torch_res = model_results["PyTorch"]
                for format_name in ["OpenVINO (FP32)", "OpenVINO (FP16)", "OpenVINO (INT8)"]:
                    if format_name in model_results:
                        result = model_results[format_name]
                        avg_speedup = torch_res["avg_time"] / result["avg_time"] if result["avg_time"] > 0 else 0
                        fps_speedup = result["fps"] / torch_res["fps"] if torch_res["fps"] > 0 else 0
                        print(f"{format_name:<20} {avg_speedup:<15.2f}x {fps_speedup:<15.2f}x")

            print("=" * 60 + "\n")

    # Print overall summary
    if all_models_results:
        print("\n" + "#" * 60)
        print("# OVERALL MULTI-MODEL BENCHMARK SUMMARY")
        print("#" * 60 + "\n")

        print("=" * 80)
        print(f"{'Model':<15} {'Format':<20} {'Avg Time (s)':<15} {'FPS':<15}")
        print("-" * 80)

        for model_name, model_results in all_models_results.items():
            for format_name in ["PyTorch", "OpenVINO (FP32)", "OpenVINO (FP16)", "OpenVINO (INT8)"]:
                if format_name in model_results:
                    result = model_results[format_name]
                    print(f"{model_name:<15} {format_name:<20} {result['avg_time']:<15.4f} {result['fps']:<15.2f}")

        print("=" * 80 + "\n")

        # Save results to Excel
        if args.output:
            output_path = args.output
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"multi_model_benchmark_results_{timestamp}.xlsx"

        config = {
            "models": models,
            "category": args.category,
            "device": f"Torch={torch_device.upper()}, OpenVINO={ov_device}",
            "num_inferences": args.num_inferences,
            "test_images": len(data.test_data),
        }

        save_results_to_excel(all_models_results, config, output_path)
    else:
        print("\nNo models were successfully benchmarked.")


if __name__ == "__main__":
    main()
