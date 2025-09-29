# Copyright (C) 2025 Intel Corporation
# SPDX-License-Identifier: Apache-2.0

"""Benchmark Patchcore model on multiple datasets with different configurations."""

import argparse
import time
from pathlib import Path

import torch
from lightning import seed_everything
from openpyxl import Workbook

from anomalib.data import Folder
from anomalib.engine import Engine, SingleXPUStrategy, XPUAccelerator
from anomalib.models import Patchcore

CUSTOM_DATASETS_ROOT = "datasets/Custom/MVTecCustom/screw"


def get_device_engine(device: str) -> Engine:
    """Create engine based on device type."""
    if device == "gpu":
        if not torch.cuda.is_available():
            msg = "CUDA is not available. Cannot use GPU strategy."
            raise RuntimeError(msg)
        print("Using GPU strategy.")
        return Engine()
    if device == "xpu":
        if not torch.xpu.is_available():
            msg = "XPU is not available. Cannot use XPU strategy."
            raise RuntimeError(msg)
        print("Using XPU strategy.")
        print(f"XPU Available: {torch.xpu.is_available()}")
        print(f"XPU Device Count: {torch.xpu.device_count()}")
        print(f"XPU Current Device: {torch.xpu.current_device()}")
        print(f"XPU Device Name: {torch.xpu.get_device_name(0)}")
        print(f"XPU Device Properties: {torch.xpu.get_device_properties(0)}")
        return Engine(strategy=SingleXPUStrategy(), accelerator=XPUAccelerator())
    if device == "cpu":
        print("Using CPU strategy.")
        return Engine(accelerator="cpu")
    msg = f"Invalid device: {device}. Choose from 'gpu', 'xpu', or 'cpu'."
    raise ValueError(msg)


def find_dataset_path(dataset_num: str, root_path: str) -> str:
    """Find the actual directory path for a given dataset number."""
    root = Path(root_path)

    if not root.exists():
        msg = f"Dataset root path {root_path} does not exist."
        raise ValueError(msg)

    # Look for directory ending with MVTeC_screw_{dataset_num}
    for item in root.iterdir():
        if item.is_dir() and item.name.endswith(f"MVTeC_screw_{dataset_num}"):
            return str(item)

    msg = f"Could not find dataset directory for screw_{dataset_num} in {root_path}"
    raise ValueError(msg)


def validate_datasets(dataset_numbers: list[str], root_path: str) -> list[str]:
    """Validate that all specified datasets exist and return valid ones."""
    valid_datasets: list[str] = []
    root = Path(root_path)

    if not root.exists():
        return valid_datasets

    # Get all existing dataset directories
    existing_datasets = {
        item.name.split("_")[-1]: item.name for item in root.iterdir() if item.is_dir() and "MVTeC_screw_" in item.name
    }

    for dataset_num in dataset_numbers:
        if dataset_num in existing_datasets:
            valid_datasets.append(dataset_num)
            print(f"✓ Found dataset: screw_{dataset_num}")
        else:
            print(f"✗ Warning: Could not find dataset directory for screw_{dataset_num} in {root_path}")

    return valid_datasets


def create_datamodule(dataset_name: str, dataset_root: str) -> Folder:
    """Create datamodule for a specific dataset."""
    # dataset_name is the number (50, 100, 150, 200)
    dataset_path = find_dataset_path(dataset_name, dataset_root)

    return Folder(
        name=f"screw_{dataset_name}",
        root=dataset_path,
        normal_dir="train/good",
        normal_test_dir="test_good",
        abnormal_dir="test",
        mask_dir="ground_truth",
    )


def run_benchmark(dataset_name: str, run_number: int, device: str, num_runs: int, pause_seconds: int = 0) -> dict:
    """Run a single benchmark for a dataset."""
    print(f"\n{'=' * 60}")
    print(f"Running {dataset_name} - Run {run_number + 1}/{num_runs}")
    print(f"{'=' * 60}")
    # Set seed for reproducibility
    seed = 42 + run_number  # Different seed for each run
    seed_everything(seed=seed, workers=True, verbose=True)

    # Create datamodule
    datamodule = create_datamodule(dataset_name, CUSTOM_DATASETS_ROOT)
    model = Patchcore()
    engine = get_device_engine(device)

    # Training
    print("Starting training...")
    train_start_time = time.time()
    engine.fit(datamodule=datamodule, model=model)
    train_end_time = time.time()
    training_time = train_end_time - train_start_time

    # Testing
    print("Starting testing...")
    test_start_time = time.time()
    results = engine.test(datamodule=datamodule, model=model)
    test_end_time = time.time()
    testing_time = test_end_time - test_start_time

    # Extract metrics
    result_dict = {
        "dataset": dataset_name,
        "run": run_number + 1,
        "device": device,
        "training_time": training_time,
        "testing_time": testing_time,
        "image_AUROC": results[0].get("image_AUROC", 0.0),
        "image_F1Score": results[0].get("image_F1Score", 0.0),
        "pixel_AUROC": results[0].get("pixel_AUROC", 0.0),
        "pixel_F1Score": results[0].get("pixel_F1Score", 0.0),
    }

    print(f"Training time: {training_time:.2f}s")
    print(f"Testing time: {testing_time:.2f}s")
    print(f"Results: {result_dict}")

    # Add pause between runs (except after the last run)
    if pause_seconds > 0:
        print(f"Pausing for {pause_seconds} seconds...")
        time.sleep(pause_seconds)

    return result_dict


def create_excel_report(all_results: list[dict], output_file: str) -> None:
    """Create Excel report with results and averages."""
    # Create output directory if it doesn't exist
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create detailed results sheet
    ws_details = wb.create_sheet("Detailed Results")
    headers = [
        "Dataset",
        "Run",
        "Device",
        "Training Time (s)",
        "Testing Time (s)",
        "Image AUROC",
        "Image F1Score",
        "Pixel AUROC",
        "Pixel F1Score",
    ]
    ws_details.append(headers)

    for result in all_results:
        row = [
            result["dataset"],
            result["run"],
            result["device"],
            result["training_time"],
            result["testing_time"],
            result["image_AUROC"],
            result["image_F1Score"],
            result["pixel_AUROC"],
            result["pixel_F1Score"],
        ]
        ws_details.append(row)

    # Create averages sheet
    ws_avg = wb.create_sheet("Averages per Dataset")
    avg_headers = [
        "Dataset",
        "Device",
        "Avg Training Time (s)",
        "Avg Testing Time (s)",
        "Avg Image AUROC",
        "Avg Image F1Score",
        "Avg Pixel AUROC",
        "Avg Pixel F1Score",
    ]
    ws_avg.append(avg_headers)

    # Calculate averages per dataset
    datasets = list({result["dataset"] for result in all_results})
    for dataset in sorted(datasets):
        dataset_results = [r for r in all_results if r["dataset"] == dataset]

        avg_row = [
            dataset,
            dataset_results[0]["device"],  # Assuming same device for all runs
            sum(r["training_time"] for r in dataset_results) / len(dataset_results),
            sum(r["testing_time"] for r in dataset_results) / len(dataset_results),
            sum(r["image_AUROC"] for r in dataset_results) / len(dataset_results),
            sum(r["image_F1Score"] for r in dataset_results) / len(dataset_results),
            sum(r["pixel_AUROC"] for r in dataset_results) / len(dataset_results),
            sum(r["pixel_F1Score"] for r in dataset_results) / len(dataset_results),
        ]
        ws_avg.append(avg_row)

    # Auto-adjust column widths
    for ws in [ws_details, ws_avg]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
                else:
                    max_length = max(max_length, 10)
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    print(f"\nResults saved to: {output_file}")


def main() -> None:
    """Run benchmark with command line arguments."""
    parser = argparse.ArgumentParser(description="Benchmark Patchcore on multiple datasets")
    parser.add_argument("--num_runs", type=int, default=5, help="Number of runs per dataset (default: 5)")
    parser.add_argument(
        "--device",
        type=str,
        choices=["gpu", "xpu", "cpu"],
        default="gpu",
        help="Device to use for training/testing (default: gpu)",
    )
    parser.add_argument(
        "--pause_seconds",
        type=int,
        default=20,
        help="Pause duration in seconds between runs (default: 20)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output Excel file name (default: auto-generated based on parameters)",
    )
    parser.add_argument(
        "--datasets",
        type=str,
        nargs="+",
        default=["50", "100", "150", "200"],
        help="Dataset numbers to run (default: 50 100 150 200)",
    )

    args = parser.parse_args()

    # Generate dynamic filename if not provided
    if args.output is None:
        dataset_suffix = "_".join(args.datasets)
        filename = f"bm_patchcore_runs-{args.num_runs}_{dataset_suffix}_{args.device}.xlsx"
        args.output = f"results/patchcore_bm/{filename}"

    print("Benchmark Configuration:")
    print(f"- Number of runs per dataset: {args.num_runs}")
    print(f"- Device: {args.device}")
    print(f"- Pause between runs: {args.pause_seconds} seconds")
    print(f"- Output file: {args.output}")
    print(f"- Datasets to run: {args.datasets}")

    # Validate datasets
    datasets = validate_datasets(args.datasets, CUSTOM_DATASETS_ROOT)
    if not datasets:
        print(f"No valid datasets found in {CUSTOM_DATASETS_ROOT}")
        return

    print(f"\nWill run benchmarks on {len(datasets)} datasets: {datasets}")

    all_results = []

    # Run benchmarks
    total_runs = len(datasets) * args.num_runs
    current_run = 0

    for dataset in datasets:
        for run in range(args.num_runs):
            current_run += 1
            is_last_run = current_run == total_runs

            try:
                # Don't pause after the very last run
                pause_time = 0 if is_last_run else args.pause_seconds
                result = run_benchmark(dataset, run, args.device, args.num_runs, pause_time)
                all_results.append(result)
            except (RuntimeError, ValueError, OSError) as e:
                print(f"Error running {dataset} run {run + 1}: {e}")
                continue

    # Create Excel report
    if all_results:
        create_excel_report(all_results, args.output)
        print(f"\nBenchmark completed! Processed {len(all_results)} total runs.")
    else:
        print("No results to save.")


if __name__ == "__main__":
    main()
